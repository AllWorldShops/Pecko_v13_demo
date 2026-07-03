import base64
import io
from odoo import models, fields, _
from odoo.exceptions import UserError


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise UserError(_("The 'openpyxl' library is required. Install it: pip install openpyxl"))

FG_PREFIXES = ('2200', '2100', '5100', 'CR00')
HEADER_BG   = 'FF1F3864'
RM_BG       = 'FFC2F1C8'
FG_BG       = 'FFC1E5F5'
TOTAL_BG_RM = 'FF388E3C'
TOTAL_BG_FG = 'FF1565C0'
THIN        = Side(style='thin')

# ── Column definitions ────────────────────────────────────────────────────────
# Non-SG: S.No … Finished Goods | Total Valuation
# SG:     S.No … Finished Goods | RMA Delivery | RMA Receipt | Total Valuation

NON_SG_HEADERS = [
    'S.No',
    'Product/Internal Reference',
    'Product/MPN/Customer/Supplier Part No',
    'Product/Description',
    'Unit of Measure',
    'Cost',
    'Quantity on hand',
    'Main warehouse',
    'Production floor',
    'Finished Goods',
    'Total Valuation',
]
NON_SG_WIDTHS = [6, 28, 28, 40, 14, 10, 16, 16, 16, 16, 18]

SG_HEADERS = [
    'S.No',
    'Product/Internal Reference',
    'Product/MPN/Customer/Supplier Part No',
    'Product/Description',
    'Unit of Measure',
    'Cost',
    'Quantity on hand',
    'Main warehouse',
    'Production floor',
    'Finished Goods',
    'RMA Delivery',
    'RMA Receipt',
    'Total Valuation',
]
SG_WIDTHS = [6, 28, 28, 40, 14, 10, 16, 16, 16, 16, 16, 16, 18]


def _border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _apply(cell, bg=None, bold=False, halign='left', wrap=False, num_fmt=None):
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    else:
        cell.fill = PatternFill('none')
    font_color = 'FFFFFFFF' if bg in (HEADER_BG, TOTAL_BG_RM, TOTAL_BG_FG) else 'FF000000'
    cell.font = Font(name='Arial', bold=bold, size=10, color=font_color)
    cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    cell.border = _border()
    if num_fmt:
        cell.number_format = num_fmt


# ── Shared sheet-building logic (used by wizard AND audit model) ──────────────

def build_workbook(env, companies=None):
    """
    Build and return (workbook_bytes, total_product_count).
    Can be called from the wizard or from the scheduled audit action.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if companies is None:
        companies = env['res.company'].search([], order='name')

    total = 0
    for company in companies:
        count = _build_company_sheet(env, wb, company, company.name[:31])
        total += count

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), total


def _is_fg(product):
    return (product.default_code or '').strip().startswith(FG_PREFIXES)


def _is_sg(company):
    return (company.country_id.code or '') == 'SG'


def _get_locations(env, company):
    Location = env['stock.location']

    def find(names, usage=None):
        domain = [('active', '=', True), ('company_id', 'in', [company.id, False])]
        if usage:
            domain.append(('usage', '=', usage))
        for name in names:
            loc = Location.search(domain + [('name', 'ilike', name)], limit=1)
            if loc:
                return loc.id
        return None

    locs = {
        'main': find(['Stock', 'Main Warehouse', 'Main RAW Material Warehouse', 'WH/Stock'],         usage='internal'),
        'prod': find(['Production', 'Production Floor', 'WIP'],       usage='internal'),
        'fg':   find(['Finished Goods', 'FG', 'Finished Product'],    usage='internal'),
    }
    if _is_sg(company):
        locs['rma_delivery'] = find(['RMA Delivery', 'RMA Out', 'Returns Out'])
        locs['rma_receipt']  = find(['RMA Receipt',  'RMA In',  'Returns In'])
    return locs


def _qty_at(env, product_id, location_id):
    if not location_id:
        return 0.0
    quants = env['stock.quant'].search([
        ('product_id', '=', product_id),
        ('location_id', 'child_of', location_id),
    ])
    return sum(quants.mapped('quantity'))


def _get_products_for_company(env, company):
    locations = env['stock.location'].search([
        ('usage', 'in', ['internal', 'transit', 'customer', 'supplier']),
        ('active', '=', True),
        ('company_id', 'in', [company.id, False]),
    ])
    quants = env['stock.quant'].search([
        ('location_id', 'in', locations.ids),
        ('quantity', '>', 0),
        ('product_id.active', '=', True),
    ])
    seen = {}
    for q in quants:
        if q.product_id.id not in seen:
            seen[q.product_id.id] = q.product_id
    return list(seen.values())


def _build_company_sheet(env, wb, company, sheet_name):
    ws  = wb.create_sheet(title=sheet_name)
    sg  = _is_sg(company)
    headers  = SG_HEADERS    if sg else NON_SG_HEADERS
    widths   = SG_WIDTHS     if sg else NON_SG_WIDTHS
    num_cols = len(headers)
    val_col  = headers.index('Total Valuation') + 1   # 1-based

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    today = fields.Date.today()
    locs  = _get_locations(env, company)

    # Rows 1–3: meta
    for r, (label, value, fmt) in enumerate([
        ('Company', company.name, None),
        ('Date',    today,        'DD MMM YYYY'),
        ('Legend',  'Green (Raw Materials), Blue (Finished Goods)', None),
    ], 1):
        ws.row_dimensions[r].height = 18
        _apply(ws.cell(r, 1, label), bold=True)
        c = ws.cell(r, 2, value)
        _apply(c)
        if fmt:
            c.number_format = fmt
        for col in range(3, num_cols + 1):
            _apply(ws.cell(r, col))

    # Row 4: blank separator
    for col in range(1, num_cols + 1):
        ws.cell(4, col).border = _border()

    # Row 5: header
    ws.row_dimensions[5].height = 30
    for ci, h in enumerate(headers, 1):
        _apply(ws.cell(5, ci, h), bg=HEADER_BG, bold=True, halign='center', wrap=True)

    products = _get_products_for_company(env, company)
    raw_list = sorted([p for p in products if not _is_fg(p)], key=lambda p: p.default_code or '')
    fg_list  = sorted([p for p in products if     _is_fg(p)], key=lambda p: p.default_code or '')

    current_row   = [6]
    serial        = [1]
    rm_valuation  = [0.0]
    fg_valuation  = [0.0]

    def write_product_row(product, bg):
        qty_main = _qty_at(env, product.id, locs.get('main'))
        qty_prod = _qty_at(env, product.id, locs.get('prod'))
        qty_fg   = _qty_at(env, product.id, locs.get('fg'))

        # Use sum of all internal quants for this company as qty_tot
        # so products with stock outside the 3 mapped locations are
        # still counted correctly, and zero-qty rows are skipped.
        all_locs = env['stock.location'].search([
            ('usage', '=', 'internal'),
            ('active', '=', True),
            ('company_id', 'in', [company.id, False]),
        ])
        all_quants = env['stock.quant'].search([
            ('product_id', '=', product.id),
            ('location_id', 'in', all_locs.ids),
            ('quantity', '>', 0),
        ])
        qty_tot = sum(all_quants.mapped('quantity'))
        if qty_tot <= 0:
            return   # skip — no actual stock in any internal location

        # qty_tot  = qty_main + qty_prod + qty_fg
        # cost      = product.standard_price or 0.0
        cost      = product.with_company(company).standard_price or 0.0
        valuation = cost * qty_tot

        # Build data list matching column order
        data = [
            serial[0],
            product.default_code or '',
            product.x_studio_field_qr3ai or '',
            product.x_studio_field_mHzKJ or '',
            product.uom_id.name if product.uom_id else '',
            cost,
            qty_tot,
            qty_main,
            qty_prod,
            qty_fg,
        ]
        if sg:
            data.append(_qty_at(env, product.id, locs.get('rma_delivery')))
            data.append(_qty_at(env, product.id, locs.get('rma_receipt')))
        data.append(valuation)   # Total Valuation — always last

        ws.row_dimensions[current_row[0]].height = 15
        for ci, val in enumerate(data, 1):
            fmt = ('#,##0.00'  if ci == val_col or ci == 6 else
                   '#,##0.###' if ci >= 7                  else None)
            _apply(ws.cell(current_row[0], ci, val), bg=bg,
                   halign='right' if ci >= 6 else 'left', num_fmt=fmt)

        if bg == RM_BG:
            rm_valuation[0] += valuation
        else:
            fg_valuation[0] += valuation

        current_row[0] += 1
        serial[0]      += 1

    def write_grand_total(label, valuation, total_bg):
        ws.row_dimensions[current_row[0]].height = 20
        for ci in range(1, num_cols + 1):
            val = label if ci == 1 else (round(valuation, 2) if ci == val_col else '')
            fmt = '#,##0.00' if ci == val_col else None
            _apply(ws.cell(current_row[0], ci, val), bg=total_bg, bold=True,
                   halign='left' if ci == 1 else 'right', num_fmt=fmt)
        current_row[0] += 1

    # Raw Materials
    for p in raw_list:
        write_product_row(p, RM_BG)
    if raw_list:
        write_grand_total('Grand Total – Raw Materials', rm_valuation[0], TOTAL_BG_RM)

    # Separator
    if raw_list and fg_list:
        for _ in range(2):
            for col in range(1, num_cols + 1):
                ws.cell(current_row[0], col).border = _border()
            current_row[0] += 1

    # Finished Goods
    for p in fg_list:
        write_product_row(p, FG_BG)
    if fg_list:
        write_grand_total('Grand Total – Finished Goods', fg_valuation[0], TOTAL_BG_FG)

    ws.freeze_panes = 'A6'
    return len(raw_list) + len(fg_list)


# ── Wizard model ──────────────────────────────────────────────────────────────

class AuditStockReportWizard(models.TransientModel):
    _name        = 'audit.stock.report.wizard'
    _description = 'Audit Monthly Stock Report Wizard'

    excel_file     = fields.Binary(string='Excel Report', readonly=True)
    excel_filename = fields.Char(readonly=True)

    def action_generate_report(self):
        self.ensure_one()
        data, count = build_workbook(self.env)
        if not count:
            raise UserError(_(
                'No stock data found in internal locations for any company.\n'
                'Please ensure stock has been received or inventory '
                'adjustments have been validated.'
            ))
        # Also save to audit log
        self.env['audit.monthly.report'].create_from_data(data)

        filename = f"stock_report_{fields.Date.today().strftime('%Y%m%d')}.xlsx"
        self.write({
            'excel_file':     base64.b64encode(data),
            'excel_filename': filename,
        })
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }
